#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
TI Stats Updater
Scrapes completed games from twilightwars.com and updates raw_data.xlsx,
then regenerates the JSON for the dashboard.
"""

import sys
import json
import requests
import subprocess
import pandas as pd
from pathlib import Path
from datetime import datetime
from playwright.sync_api import sync_playwright

BASE = "https://www.twilightwars.com"
REPO = Path(__file__).parent
EXCEL = REPO / "data" / "raw_data.xlsx"

# twilightwars username → real player name
USERNAME_MAP = {
    "ManolosMagnos": "Manu",
    "Thomas": "Thomas",
    "Keineui": "Eric",
    "L1z4Rd": "Frank",
    "Starki": "Starki",
}

# Full faction name → short name (from factions sheet, plus extras)
FACTION_MAP = {
    "The Arborec": "Arborec",
    "The Barony of Letnev": "Barony",
    "The Clan of Saar": "Saar",
    "The Embers of Muaat": "Muaat",
    "The Emirates of Hacan": "Hacan",
    "The Empyrean": "Empyrean",
    "The Federation of Sol": "Sol",
    "The Ghosts of Creuss": "Creuss",
    "The L1Z1X Mindset": "L1Z1X",
    "The Mentak Coalition": "Mentak",
    "The Naalu Collective": "Naalu",
    "The Nekro Virus": "Nekro",
    "Sardakk N'orr": "Sardakk",
    "The Sardakk N'orr": "Sardakk",
    "The Universities of Jol-Nar": "Jol",
    "The Winnu": "Winnu",
    "The Xxcha Kingdom": "Xxcha",
    "The Yssaril Tribes": "Yssaril",
    # PoK factions
    "The Argent Flight": "Argent",
    "The Vuil'Raith Cabal": "Vuil'Raith",
    "The Mahact Gene-Sorcerers": "Mahact",
    "The Nomad": "Nomad",
    "The Naaz-Rokha Alliance": "Naaz",
    "The Titans of Ul": "Titans",
    "The Empyrean": "Empyrean",
    "The Keleres": "Keleres",
}

KNOWN_PLAYERS = set(USERNAME_MAP.keys())

# Games manually excluded from the dashboard (e.g. doesn't count for the group)
EXCLUDED_GAME_IDS = {
    "69256c61fa804c7cd64e3d0f",  # Die lustigen 5 - marked irrelevant by Manu
}


def login():
    """Login and return a requests session with auth cookies."""
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()
        page.goto(f"{BASE}/login", wait_until="networkidle")
        page.fill("input[name=email]", "steiner-manuel@web.de")
        page.fill("input[name=password]", "a424424424")
        page.click("button[type=submit]")
        page.wait_for_load_state("networkidle")
        cookies = {c["name"]: c["value"] for c in page.context.cookies()}

        # Get profile game links while we're here
        page.goto(f"{BASE}/u/me", wait_until="networkidle")
        links = page.query_selector_all("a")
        game_ids = []
        for link in links:
            href = link.get_attribute("href") or ""
            if href.startswith("/games/"):
                gid = href.replace("/games/", "").strip()
                if gid and gid not in game_ids:
                    game_ids.append(gid)

        browser.close()

    session = requests.Session()
    session.cookies.update(cookies)
    return session, game_ids


def derive_win_fields(summary, players_data, log):
    """Auto-detect win_category and winner_had_support from API data."""
    win_category = "Aktionsphase" if summary.get("phase") == "Action" else "Statusphase"

    winner_player_doc_id = summary.get("winner")
    if not winner_player_doc_id:
        return win_category, 0

    winner = next((p for p in players_data if p["_id"] == winner_player_doc_id), None)
    if not winner:
        return win_category, 0

    winner_user_id = winner["user"]["_id"]
    total_vp = winner.get("victoryPoints", 0)
    counted = 0

    # Public objectives
    for obj in summary.get("publicObjectives", []):
        for score in obj.get("scores", []):
            if score["playerId"] == winner_player_doc_id:
                counted += obj["victoryPoints"]

    # Secret objectives
    for obj in winner.get("revealedSecretObjectives", []):
        for score in obj.get("scores", []):
            if score["playerId"] == winner_player_doc_id:
                counted += obj["victoryPoints"]

    # Custodians token
    if winner.get("hasCustodiansToken"):
        counted += 1

    # Imperial strategy card: +1 VP if winner controlled Mecatol Rex when they played it
    log_chron = list(reversed(log))
    mecatol_holder_user_id = None
    for event in log_chron:
        ev = event.get("event", "")
        user = event.get("user")
        details = event.get("details", {})
        if ev == "established control" and "Mecatol Rex" in (details.get("planetNames") or []):
            mecatol_holder_user_id = user
        if (ev == "primary ability resolved"
                and details.get("strategyCard") == "Imperial"
                and user == winner_user_id
                and mecatol_holder_user_id == winner_user_id):
            counted += 1

    gap = total_vp - counted
    winner_had_support = 1 if gap > 0 else 0
    return win_category, winner_had_support


def get_game_data(session, game_id):
    """Fetch summary + players for a game. Returns None if not a group game."""
    summary = session.get(f"{BASE}/games/{game_id}/summary").json()
    players_data = session.get(f"{BASE}/games/{game_id}/players").json()
    log = session.get(f"{BASE}/games/{game_id}/log").json()

    # Only process games with at least one known friend-group player (besides Manu)
    usernames = {p["user"]["username"] for p in players_data}
    group_players = usernames & KNOWN_PLAYERS
    if len(group_players) < 2:
        return None  # Not a group game

    # Parse dates
    created = datetime.fromtimestamp(summary["createdAt"]).date()
    completed = datetime.fromtimestamp(summary.get("completedAt", summary["updatedAt"])).date()

    win_category, winner_had_support = derive_win_fields(summary, players_data, log)

    game_row = {
        "game_id": game_id,
        "game_medium": "online",
        "game_name": summary["name"],
        "game_max_victory_points": summary["victoryPoints"],
        "start_date": created,
        "end_date": completed,
        "rounds": summary["round"],
        "win_category": win_category,
        "winner_had_support": winner_had_support,
        "win_description": None,  # optional, filled by prompt
    }

    result_rows = []
    for p in players_data:
        username = p["user"]["username"]
        player_name = USERNAME_MAP.get(username, username)
        faction_full = p.get("faction", "")
        faction_short = FACTION_MAP.get(faction_full)
        if not faction_short:
            print(f"  WARNING: Unknown faction '{faction_full}' for {username}. Add to FACTION_MAP.")
            faction_short = faction_full  # fallback

        result_rows.append({
            "game_id": game_id,
            "player_name": player_name,
            "victory_points": p.get("victoryPoints"),
            "faction_short_name": faction_short,
            "starting_position": p.get("number"),
        })

    return game_row, result_rows, players_data


def prompt_manual_fields(game_row, players_data):
    """Ask for the optional win_description."""
    print(f"\n--- {game_row['game_name']} ---")
    print(f"    Players: {', '.join(p['user']['username'] + ' (' + str(p.get('victoryPoints', '?')) + ' VP)' for p in players_data)}")
    print(f"    Auto-detected: win_category={game_row['win_category']}, winner_had_support={game_row['winner_had_support']}")

    desc = input("  Win description (optional, press Enter to skip): ").strip()
    game_row["win_description"] = desc if desc else None
    return game_row


def update_excel(new_games, new_results):
    """Append new rows to the Excel file."""
    from openpyxl import load_workbook

    df_games_existing = pd.read_excel(EXCEL, sheet_name="games")
    df_results_existing = pd.read_excel(EXCEL, sheet_name="results")

    df_games_updated = pd.concat([df_games_existing, pd.DataFrame(new_games)], ignore_index=True)
    df_results_updated = pd.concat([df_results_existing, pd.DataFrame(new_results)], ignore_index=True)

    wb = load_workbook(EXCEL)

    def write_df_to_sheet(df, sheet_name):
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None
        for c_idx, col_name in enumerate(df.columns, start=1):
            ws.cell(row=1, column=c_idx, value=col_name)
        for r_idx, row in enumerate(df.itertuples(index=False), start=2):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)

    write_df_to_sheet(df_games_updated, "games")
    write_df_to_sheet(df_results_updated, "results")

    wb.save(EXCEL)
    print(f"[OK] Excel updated: {len(new_games)} new game(s), {len(new_results)} new result row(s)")


def main(game_id_arg=None):
    print("Logging in to twilightwars.com...")
    session, profile_game_ids = login()
    print(f"Found {len(profile_game_ids)} games on profile.")

    # Load existing game IDs
    df_existing = pd.read_excel(EXCEL, sheet_name="games")
    existing_ids = set(df_existing["game_id"].astype(str))

    # Determine which games to process
    if game_id_arg:
        to_process = [game_id_arg]
    else:
        to_process = [gid for gid in profile_game_ids if gid not in existing_ids and gid not in EXCLUDED_GAME_IDS]

    if not to_process:
        print("No new games found. Dashboard is up to date!")
        return

    print(f"\nFound {len(to_process)} new game(s) to process.")

    new_games = []
    new_results = []

    for game_id in to_process:
        print(f"\nFetching game {game_id}...")
        result = get_game_data(session, game_id)
        if result is None:
            print(f"  Skipping (not a group game).")
            continue

        game_row, result_rows, players_data = result
        print(f"  Game: {game_row['game_name']}")
        print(f"  Date: {game_row['start_date']} → {game_row['end_date']}")
        print(f"  Players: {', '.join(r['player_name'] + ' (' + str(r['victory_points']) + ' VP)' for r in result_rows)}")

        game_row = prompt_manual_fields(game_row, players_data)

        new_games.append(game_row)
        new_results.extend(result_rows)

    if not new_games:
        print("\nNo group games to add.")
        return

    print("\nUpdating Excel...")
    update_excel(new_games, new_results)

    print("\nRegenerating dashboard JSON...")
    subprocess.run(
        [sys.executable, str(REPO / "process_data_for_web.py")],
        cwd=str(REPO),
        check=True
    )

    print("\nCommitting...")
    git_env = {
        **__import__("os").environ,
        "GIT_AUTHOR_NAME": "Manuel Rademaker",
        "GIT_AUTHOR_EMAIL": "manuel-rademaker@outlook.de",
        "GIT_COMMITTER_NAME": "Manuel Rademaker",
        "GIT_COMMITTER_EMAIL": "manuel-rademaker@outlook.de",
    }
    subprocess.run(["git", "add", "data/raw_data.xlsx", "website/public/data/ti_data.json"],
                   cwd=str(REPO), check=True, env=git_env)
    names = ", ".join(g["game_name"] for g in new_games)
    subprocess.run(
        ["git", "commit", "-m", f"Add game data: {names}"],
        cwd=str(REPO), check=True, env=git_env
    )
    subprocess.run(["git", "push"], cwd=str(REPO), check=True, env=git_env)
    print(f"\nDone! Dashboard will update on Netlify shortly.")


if __name__ == "__main__":
    game_id = sys.argv[1] if len(sys.argv) > 1 else None
    main(game_id)
